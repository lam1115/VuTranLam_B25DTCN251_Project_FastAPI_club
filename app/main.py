from fastapi import FastAPI, Request
from slowapi.errors import RateLimitExceeded
from slowapi import _rate_limit_exceeded_handler
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

from app.db.database import Base, engine
from app.routers.authentication_router import router as authentication
from app.routers.user_router import router as user
from app.routers.club_router import router as club
from app.routers.ClubMember_router import router as member
from app.routers.activity_club_router import router as activity_club
from app.core.limiter import limit


from app.core.exception_handlers import AppException, app_exception_handler

Base.metadata.create_all(bind=engine)

app = FastAPI(title="STUDENT CLUB MANAGEMENT API")

app.state.limiter = limit
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.add_exception_handler(AppException, app_exception_handler)


app.include_router(authentication)
app.include_router(user)
app.include_router(club)
app.include_router(member)
app.include_router(activity_club)


# Health check endpoint
@app.get("/health", tags=["Health Check"])
def health_check():
    return {"status": "ok", "message": "Service is running healthy"}


# EXCEL_FILE = "API_Test_Checklist_Student_Club.xlsx"


# def log_to_excel(method: str, path: str, status_code: int, detail: str):
#     # Khởi tạo file nếu chưa tồn tại
#     if not os.path.exists(EXCEL_FILE):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "API Test Log"
#         ws.append(
#             ["Thời gian", "Method", "Endpoint", "HTTP Code", "Ghi chú / Response"]
#         )
#         wb.save(EXCEL_FILE)

#     # Mở và ghi log mới
#     wb = openpyxl.load_workbook(EXCEL_FILE)
#     ws = wb["API Test Log"] if "API Test Log" in wb.sheetnames else wb.active

#     ws.append(
#         [
#             datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
#             method,
#             path,
#             status_code,
#             detail,
#         ]
#     )
#     wb.save(EXCEL_FILE)


# @app.middleware("http")
# async def excel_logging_middleware(request: Request, call_next):
#     response = await call_next(request)

#     # Chỉ ghi log các API nghiệp vụ (bỏ qua các request lấy file tĩnh/docs của Swagger)
#     if not request.url.path.startswith(
#         ("/docs", "/openapi.json", "/favicon.ico", "/redoc")
#     ):
#         log_to_excel(
#             method=request.method,
#             path=request.url.path,
#             status_code=response.status_code,
#             detail=f"Tested via Swagger/Client - Status {response.status_code}",
#         )

#     return response
